import os
from openpyxl import load_workbook
from openpyxl.cell import read_only
from openpyxl.utils.exceptions import InvalidFileException
from util_tools.logs_util.recordlog import logs

class ExcelDataReader:
    """
    读取excel文件数据
    """
    def __init__(self,file_path):
        self.file_path = file_path
        abspath = os.path.abspath(self.file_path)
        try:
            if not os.path.exists(abspath):
                raise FileExistsError(f'文件路径不存在 - {abspath}')
            # read_only：控制工作簿是否只读，如果为True，工作簿以只读模式打开，不允许对工作簿修改
            self.workbook = load_workbook(filename=self.file_path,read_only=False)
            logs.info(f'excel文件打开成功【文件路径：{self.file_path}】')
        except Exception as e:
            logs.error(f'excel文件打开异常【原因：{e}】')



    def read_row(self,sheet_name="Sheet1",row_index=1):
        """
        获取excel文件的一整行
        :param sheet_name: 表名称（一般在excel表的左下角）
        :param row_index: 第几行，索引从1开始
        :return:
        """
        try:
            sheet = self.workbook[sheet_name]
            rows_data = [cell.value for cell in sheet[row_index]]
            logs.info(f'excel文件读取单行数据成功【工作簿：{sheet_name}，读取行：{row_index}】')
            return rows_data
        except Exception as e:
            logs.error(f'excel文件读取单行数据异常【原因：{e}】')
        finally:
            self.workbook.close()
            logs.info(f'excel文件关闭成功【文件路径：{self.file_path}】')


    def read_col(self,sheet_name="Sheet1",col_index=1):
        """
        获取excel文件的一整列
        :param sheet_name: 表名称
        :param col_index: 第几列，索引从1开始
        :return:
        """
        col_data = []
        try:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=sheet.min_row,max_row=sheet.max_row,min_col=sheet.min_column,max_col=col_index):
                col_data.append(row[col_index-1].value)
            logs.info(f'excel文件读取单列数据成功【工作簿：{sheet_name}，读取列：{col_index}】')
            return col_data
        except Exception as e:
            logs.error(f'excel文件读取单列数据异常【原因：{e}】')
        finally:
            self.workbook.close()
            logs.info(f'excel文件关闭成功【文件路径：{self.file_path}】')

    def read_cell_value(self,sheet_name="Sheet1",col_index=1,row_index=1):
        """
        获取单个单元格数据
        :param sheet_name: 工作簿
        :param col_index: 第几列
        :param row_index: 第几行
        :return:
        """
        try:
            sheet = self.workbook[sheet_name]
            logs.info(f'excel文件读取单个单元格数据成功【工作簿：{sheet_name}，读取行：{row_index}，读取列：{col_index}】')
            return sheet.cell(row=row_index, column=col_index).value
        except Exception as e:
            logs.error(f'excel文件读取单个单元格数据成功【原因：{e}】')
        finally:
            self.workbook.close()
            logs.info(f'excel文件关闭成功【文件路径：{self.file_path}】')

    def read_all_row(self,sheet_name="Sheet1"):
        """
        获取excel多行数据，每一行数据组成一个列表
        :param sheet_name: sheet页名称
        :return: 返回列表格式[["admin","password"],[...,...],...]
        """
        all_rows_data = []
        try:
            sheet = self.workbook[sheet_name]
            max_col = sheet.max_column
            max_row = sheet.max_row
            for row in sheet.iter_rows(min_row=1,max_row=max_row,min_col=1,max_col=max_col):
                row_data = [cell.value for cell in row]
                all_rows_data.append(row_data)
            logs.info(f'excel文件读取所有行数据成功【工作簿：{sheet_name}】')
            return all_rows_data
        except Exception as e:
            logs.error(f'excel文件读取所有行数据异常【原因：{e}】')
        finally:
            self.workbook.close()
            logs.info(f'excel文件关闭成功【文件路径：{self.file_path}】')







if __name__ == '__main__':
    exl = ExcelDataReader('../../data/login_testdata.xlsx')
    print(exl.read_col(sheet_name="Sheet1",col_index=1))
    print(exl.read_cell_value(sheet_name="Sheet1",col_index=2,row_index=2))




